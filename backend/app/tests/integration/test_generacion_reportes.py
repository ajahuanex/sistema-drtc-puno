"""
Integration Test: Generación de Reportes
Tests report generation and statistics
"""
import pytest
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from unittest.mock import patch

from app.models.mesa_partes.documento import Documento, EstadoDocumentoEnum


@pytest.mark.asyncio
class TestGeneracionReportes:
    """Test del sistema de generación de reportes"""
    
    async def test_obtener_estadisticas_generales(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user,
        test_documentos_multiples
    ):
        """
        Test de obtención de estadísticas generales
        """
        
        # Obtener estadísticas
        response = await async_client.get(
            "/api/v1/mesa-partes/reportes/estadisticas",
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        estadisticas = response.json()
        
        # Verificar estructura de estadísticas
        assert "total_documentos" in estadisticas
        assert "documentos_registrados" in estadisticas
        assert "documentos_en_proceso" in estadisticas
        assert "documentos_atendidos" in estadisticas
        assert "documentos_archivados" in estadisticas
        assert "documentos_vencidos" in estadisticas
        
        # Verificar que los totales son correctos
        assert estadisticas["total_documentos"] > 0
        assert (
            estadisticas["documentos_registrados"] +
            estadisticas["documentos_en_proceso"] +
            estadisticas["documentos_atendidos"] +
            estadisticas["documentos_archivados"]
        ) == estadisticas["total_documentos"]
        
    async def test_estadisticas_por_rango_fechas(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user
    ):
        """
        Test de estadísticas filtradas por rango de fechas
        """
        
        fecha_inicio = (datetime.utcnow() - timedelta(days=30)).isoformat()
        fecha_fin = datetime.utcnow().isoformat()
        
        response = await async_client.get(
            f"/api/v1/mesa-partes/reportes/estadisticas?fecha_inicio={fecha_inicio}&fecha_fin={fecha_fin}",
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        estadisticas = response.json()
        
        assert "periodo" in estadisticas
        assert estadisticas["periodo"]["fecha_inicio"] == fecha_inicio
        assert estadisticas["periodo"]["fecha_fin"] == fecha_fin
        
    async def test_estadisticas_por_tipo_documento(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user,
        test_documentos_multiples
    ):
        """
        Test de distribución de documentos por tipo
        """
        
        response = await async_client.get(
            "/api/v1/mesa-partes/reportes/estadisticas/por-tipo",
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        distribucion = response.json()
        
        # Verificar estructura
        assert isinstance(distribucion, list)
        
        for item in distribucion:
            assert "tipo_documento" in item
            assert "cantidad" in item
            assert "porcentaje" in item
            assert item["cantidad"] > 0
            assert 0 <= item["porcentaje"] <= 100
            
    async def test_estadisticas_por_area(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user,
        test_documentos_multiples
    ):
        """
        Test de distribución de documentos por área
        """
        
        response = await async_client.get(
            "/api/v1/mesa-partes/reportes/estadisticas/por-area",
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        distribucion = response.json()
        
        assert isinstance(distribucion, list)
        
        for item in distribucion:
            assert "area" in item
            assert "documentos_recibidos" in item
            assert "documentos_atendidos" in item
            assert "documentos_pendientes" in item
            assert "tiempo_promedio_atencion" in item
            
    async def test_metricas_tiempo_atencion(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user,
        test_documentos_atendidos
    ):
        """
        Test de métricas de tiempo de atención
        """
        
        response = await async_client.get(
            "/api/v1/mesa-partes/reportes/metricas/tiempo-atencion",
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        metricas = response.json()
        
        # Verificar métricas calculadas
        assert "tiempo_promedio_horas" in metricas
        assert "tiempo_minimo_horas" in metricas
        assert "tiempo_maximo_horas" in metricas
        assert "mediana_horas" in metricas
        assert "documentos_analizados" in metricas
        
        assert metricas["tiempo_promedio_horas"] > 0
        assert metricas["tiempo_minimo_horas"] <= metricas["tiempo_promedio_horas"]
        assert metricas["tiempo_maximo_horas"] >= metricas["tiempo_promedio_horas"]
        
    async def test_generar_reporte_personalizado(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user
    ):
        """
        Test de generación de reporte personalizado
        """
        
        reporte_config = {
            "titulo": "Reporte Mensual de Mesa de Partes",
            "fecha_inicio": (datetime.utcnow() - timedelta(days=30)).isoformat(),
            "fecha_fin": datetime.utcnow().isoformat(),
            "incluir_estadisticas": True,
            "incluir_graficos": True,
            "incluir_detalle_documentos": True,
            "filtros": {
                "estados": ["REGISTRADO", "EN_PROCESO", "ATENDIDO"],
                "prioridades": ["ALTA", "URGENTE"]
            }
        }
        
        response = await async_client.post(
            "/api/v1/mesa-partes/reportes/generar",
            json=reporte_config,
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        reporte = response.json()
        
        # Verificar estructura del reporte
        assert "id" in reporte
        assert "titulo" in reporte
        assert "fecha_generacion" in reporte
        assert "estadisticas" in reporte
        assert "documentos" in reporte
        
    async def test_exportar_reporte_excel(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user,
        test_documentos_multiples
    ):
        """
        Test de exportación de reporte a Excel
        """
        
        filtros = {
            "fecha_inicio": (datetime.utcnow() - timedelta(days=30)).isoformat(),
            "fecha_fin": datetime.utcnow().isoformat()
        }
        
        response = await async_client.get(
            "/api/v1/mesa-partes/reportes/exportar/excel",
            params=filtros,
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        # Verificar que el contenido es un archivo Excel válido
        excel_content = BytesIO(response.content)
        workbook = openpyxl.load_workbook(excel_content)
        
        # Verificar hojas del libro
        assert "Resumen" in workbook.sheetnames
        assert "Documentos" in workbook.sheetnames
        
        # Verificar contenido de la hoja de documentos
        sheet_documentos = workbook["Documentos"]
        assert sheet_documentos["A1"].value == "Número Expediente"
        assert sheet_documentos["B1"].value == "Tipo Documento"
        assert sheet_documentos["C1"].value == "Remitente"
        
    async def test_exportar_reporte_pdf(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user,
        test_documentos_multiples
    ):
        """
        Test de exportación de reporte a PDF
        """
        
        filtros = {
            "fecha_inicio": (datetime.utcnow() - timedelta(days=30)).isoformat(),
            "fecha_fin": datetime.utcnow().isoformat()
        }
        
        response = await async_client.get(
            "/api/v1/mesa-partes/reportes/exportar/pdf",
            params=filtros,
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        
        # Verificar que el contenido es un PDF válido
        pdf_content = response.content
        assert pdf_content.startswith(b"%PDF")
        
    async def test_grafico_tendencias_temporal(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user,
        test_documentos_multiples
    ):
        """
        Test de generación de datos para gráfico de tendencias
        """
        
        response = await async_client.get(
            "/api/v1/mesa-partes/reportes/graficos/tendencias",
            params={
                "fecha_inicio": (datetime.utcnow() - timedelta(days=30)).isoformat(),
                "fecha_fin": datetime.utcnow().isoformat(),
                "agrupacion": "dia"
            },
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        datos_grafico = response.json()
        
        # Verificar estructura de datos para gráfico
        assert "labels" in datos_grafico
        assert "datasets" in datos_grafico
        
        assert isinstance(datos_grafico["labels"], list)
        assert isinstance(datos_grafico["datasets"], list)
        
        for dataset in datos_grafico["datasets"]:
            assert "label" in dataset
            assert "data" in dataset
            assert len(dataset["data"]) == len(datos_grafico["labels"])
            
    async def test_reporte_documentos_vencidos(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user,
        test_documentos_vencidos
    ):
        """
        Test de reporte de documentos vencidos
        """
        
        response = await async_client.get(
            "/api/v1/mesa-partes/reportes/documentos-vencidos",
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        reporte = response.json()
        
        assert "total_vencidos" in reporte
        assert "documentos" in reporte
        
        # Verificar que todos los documentos están vencidos
        for doc in reporte["documentos"]:
            assert doc["fecha_limite"] is not None
            fecha_limite = datetime.fromisoformat(doc["fecha_limite"].replace("Z", "+00:00"))
            assert fecha_limite < datetime.utcnow()
            
    async def test_reporte_por_usuario(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user,
        test_documentos_multiples
    ):
        """
        Test de reporte de productividad por usuario
        """
        
        response = await async_client.get(
            "/api/v1/mesa-partes/reportes/por-usuario",
            params={
                "fecha_inicio": (datetime.utcnow() - timedelta(days=30)).isoformat(),
                "fecha_fin": datetime.utcnow().isoformat()
            },
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        reporte = response.json()
        
        assert isinstance(reporte, list)
        
        for usuario_stats in reporte:
            assert "usuario" in usuario_stats
            assert "documentos_registrados" in usuario_stats
            assert "documentos_atendidos" in usuario_stats
            assert "derivaciones_realizadas" in usuario_stats
            
    async def test_dashboard_indicadores_clave(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user,
        test_documentos_multiples
    ):
        """
        Test de obtención de indicadores clave para dashboard
        """
        
        response = await async_client.get(
            "/api/v1/mesa-partes/reportes/dashboard",
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        dashboard = response.json()
        
        # Verificar indicadores clave
        assert "total_documentos_hoy" in dashboard
        assert "documentos_pendientes" in dashboard
        assert "documentos_urgentes" in dashboard
        assert "documentos_proximos_vencer" in dashboard
        assert "tiempo_promedio_atencion_horas" in dashboard
        assert "tasa_atencion_oportuna" in dashboard
        
        # Verificar gráficos
        assert "grafico_tendencias" in dashboard
        assert "grafico_por_tipo" in dashboard
        assert "grafico_por_area" in dashboard
        
    async def test_reporte_comparativo_periodos(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user
    ):
        """
        Test de reporte comparativo entre períodos
        """
        
        response = await async_client.get(
            "/api/v1/mesa-partes/reportes/comparativo",
            params={
                "periodo1_inicio": (datetime.utcnow() - timedelta(days=60)).isoformat(),
                "periodo1_fin": (datetime.utcnow() - timedelta(days=30)).isoformat(),
                "periodo2_inicio": (datetime.utcnow() - timedelta(days=30)).isoformat(),
                "periodo2_fin": datetime.utcnow().isoformat()
            },
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 200
        comparativo = response.json()
        
        # Verificar estructura comparativa
        assert "periodo1" in comparativo
        assert "periodo2" in comparativo
        assert "variacion" in comparativo
        
        # Verificar métricas comparadas
        assert "total_documentos" in comparativo["periodo1"]
        assert "total_documentos" in comparativo["periodo2"]
        assert "total_documentos" in comparativo["variacion"]
        
    async def test_programar_reporte_automatico(
        self,
        async_client: AsyncClient,
        db_session: AsyncSession,
        test_user
    ):
        """
        Test de programación de reporte automático
        """
        
        programacion = {
            "nombre": "Reporte Semanal",
            "frecuencia": "semanal",
            "dia_semana": "lunes",
            "hora": "08:00",
            "destinatarios": ["jefe@institucion.gob.pe"],
            "formato": "pdf",
            "incluir_estadisticas": True
        }
        
        response = await async_client.post(
            "/api/v1/mesa-partes/reportes/programar",
            json=programacion,
            headers={"Authorization": f"Bearer {test_user.token}"}
        )
        
        assert response.status_code == 201
        reporte_programado = response.json()
        
        assert reporte_programado["nombre"] == "Reporte Semanal"
        assert reporte_programado["frecuencia"] == "semanal"
        assert reporte_programado["activo"] is True
