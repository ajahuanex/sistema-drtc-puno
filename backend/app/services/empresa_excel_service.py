"""
Servicio para carga masiva de empresas desde archivos Excel - SIN DATOS MOCK
"""
import pandas as pd
import re
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional
from io import BytesIO
from app.models.empresa import (
    EmpresaCreate, 
    RazonSocial, 
    RepresentanteLegal, 
    EstadoEmpresa,
    DocumentoEmpresa,
    TipoDocumento
)
from app.services.empresa_service import EmpresaService
from app.services.configuracion_service import ConfiguracionService
from app.dependencies.db import get_database

class EmpresaExcelService:
    def __init__(self):
        self.empresa_service = None
        self.configuracion_service = None
    
    def _limpiar_valor_enum(self, valor: str) -> str:
        """Limpiar valores de enums removiendo prefijos y formateando"""
        if not isinstance(valor, str) or not valor:
            return valor
        
        # Lista de prefijos conocidos a remover
        prefijos_a_remover = [
            'EstadoEmpresa.',
            'TipoServicio.',
            'TipoEmpresa.',
            'TipoDocumento.'
        ]
        
        # Remover cualquier prefijo conocido
        for prefijo in prefijos_a_remover:
            if prefijo in valor:
                valor = valor.replace(prefijo, '')
                break
        
        # Si aún tiene punto, tomar solo la parte después del último punto
        if '.' in valor:
            valor = valor.split('.')[-1]
        
        # Reemplazar guiones bajos con espacios y capitalizar
        valor = valor.replace('_', ' ').title()
        
        return valor
        
    async def _get_empresa_service(self):
        """Obtener servicio de empresas"""
        if not self.empresa_service:
            try:
                db = await get_database()
                if db is None:
                    return None
                self.empresa_service = EmpresaService(db)
            except Exception as e:
                print(f"Error obteniendo servicio de empresas: {e}")
                return None
        return self.empresa_service
    
    async def _get_configuracion_service(self):
        """Obtener servicio de configuraciones"""
        if not self.configuracion_service:
            try:
                db = await get_database()
                if db is None:
                    return None
                self.configuracion_service = ConfiguracionService(db)
            except Exception as e:
                print(f"Error obteniendo servicio de configuraciones: {e}")
                return None
        return self.configuracion_service
        
    def generar_plantilla_excel(self) -> BytesIO:
        """Generar plantilla Excel profesional para carga masiva de empresas con múltiples hojas"""
        
        # Crear archivo Excel en memoria
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            
            # ========================================
            # HOJA 1: DATOS (donde van los datos)
            # ========================================
            # Orden específico solicitado por el usuario
            datos_vacios = {
                'RUC': ['', ''],
                'Razón Social Principal': ['', ''],
                'Dirección Fiscal': ['', ''],
                'Teléfono Contacto': ['', ''],
                'Email Contacto': ['', ''],
                'Nombres Representante': ['', ''],
                'Apellidos Representante': ['', ''],
                'DNI Representante': ['', ''],
                'Partida Registral': ['', ''],
                'Razón Social SUNAT': ['', ''],
                'Razón Social Mínimo': ['', ''],
                'Estado': ['', ''],
                'Estado SUNAT': ['', ''],
                'Tipo de Servicio': ['', ''],
                'Observaciones': ['', '']
            }
            
            df_datos = pd.DataFrame(datos_vacios)
            df_datos.to_excel(writer, sheet_name='DATOS', index=False)
            
            # ========================================
            # HOJA 2: INSTRUCCIONES
            # ========================================
            instrucciones = [
                ['INSTRUCCIONES PARA CARGA MASIVA DE EMPRESAS'],
                [''],
                ['1. CÓMO USAR ESTA PLANTILLA:'],
                ['   • Complete los datos en la hoja "DATOS"'],
                ['   • Solo RUC y Razón Social Principal son OBLIGATORIOS'],
                ['   • Todos los demás campos son OPCIONALES'],
                ['   • Vea ejemplos en la hoja "EJEMPLOS"'],
                ['   • Guarde el archivo y súbalo al sistema'],
                [''],
                ['2. CAMPOS OBLIGATORIOS:'],
                ['   • RUC: Exactamente 11 dígitos'],
                ['   • Razón Social Principal: Nombre de la empresa'],
                [''],
                ['3. CAMPOS OPCIONALES:'],
                ['   • Dirección Fiscal: Se puede completar después'],
                ['   • Teléfono Contacto: Múltiples números separar con espacios'],
                ['   • Email Contacto: Formato válido'],
                ['   • Representante: DNI, nombres y apellidos opcionales'],
                ['   • Partida Registral: Se puede completar después'],
                ['   • Estados y observaciones: Se pueden agregar después'],
                [''],
                ['4. VALIDACIONES:'],
                ['   • RUC: Exactamente 11 dígitos'],
                ['   • DNI: Máximo 8 dígitos (se completa con ceros a la izquierda)'],
                ['   • Partida Registral: Entre 1 y 9 dígitos (se completa con ceros hasta 8)'],
                ['   • Emails: Formato válido (ejemplo@dominio.com)'],
                ['   • Teléfonos: Solo números, espacios, guiones y paréntesis'],
                ['   • Múltiples teléfonos: Separar con espacios (se convertirán a comas)'],
                ['     Ejemplo: "051-123456 054-987654" se convierte en "051-123456, 054-987654"'],
                [''],
                ['5. ESTADOS VÁLIDOS:'],
                ['   • AUTORIZADA'],
                ['   • EN_TRAMITE'],
                ['   • SUSPENDIDA'],
                ['   • CANCELADA'],
                ['   • DADA_DE_BAJA'],
                [''],
                ['6. CONSEJOS:'],
                ['   • Mínimo requerido: RUC + Razón Social Principal'],
                ['   • Los campos vacíos se pueden completar después'],
                ['   • Use el formato correcto para cada tipo de dato'],
                ['   • No deje espacios al inicio o final de los datos'],
                ['   • Revise los ejemplos antes de completar'],
                [''],
                ['7. SOPORTE:'],
                ['   • Si tiene dudas, consulte con el administrador del sistema'],
                ['   • Los errores se mostrarán detalladamente en el resultado']
            ]
            
            df_instrucciones = pd.DataFrame(instrucciones, columns=['Instrucciones'])
            df_instrucciones.to_excel(writer, sheet_name='INSTRUCCIONES', index=False)
            
            # ========================================
            # HOJA 3: CAMPOS (descripción de campos)
            # ========================================
            campos_info = [
                ['Campo', 'Obligatorio', 'Tipo de Dato', 'Descripción', 'Ejemplo'],
                ['RUC', 'SÍ', 'Numérico (11 dígitos)', 'Registro Único de Contribuyente', '20123456789'],
                ['Razón Social Principal', 'SÍ', 'Texto', 'Nombre principal de la empresa', 'TRANSPORTES PUNO S.A.C.'],
                ['Dirección Fiscal', 'NO', 'Texto', 'Dirección fiscal registrada', 'AV. EJERCITO 123, PUNO'],
                ['Teléfono Contacto', 'NO', 'Texto', 'Teléfono(s) de contacto. Múltiples números separar con espacios', '051-123456 054-987654'],
                ['Email Contacto', 'NO', 'Email', 'Email de contacto de la empresa', 'contacto@empresa.com'],
                ['Nombres Representante', 'NO', 'Texto', 'Nombres del representante', 'JUAN CARLOS'],
                ['Apellidos Representante', 'NO', 'Texto', 'Apellidos del representante', 'MAMANI QUISPE'],
                ['DNI Representante', 'NO', 'Numérico (máx 8 dígitos)', 'DNI del representante legal (se completa con ceros)', '12345678'],
                ['Partida Registral', 'NO', 'Numérico (8-9 dígitos)', 'Número de partida registral (se completa con ceros)', '00123456'],
                ['Razón Social SUNAT', 'NO', 'Texto', 'Razón social según SUNAT', 'TRANSPORTES PUNO SOCIEDAD ANONIMA CERRADA'],
                ['Razón Social Mínimo', 'NO', 'Texto', 'Nombre corto de la empresa', 'TRANSPORTES PUNO'],
                ['Estado', 'NO', 'Lista', 'Estado de la empresa', 'AUTORIZADA'],
                ['Estado SUNAT', 'NO', 'Texto', 'Estado según SUNAT', 'ACTIVO'],
                ['Tipo de Servicio', 'NO', 'Lista', 'Tipo de servicio que ofrece la empresa', 'PERSONAS'],
                ['Observaciones', 'NO', 'Texto', 'Comentarios adicionales', 'Empresa especializada en transporte']
            ]
            
            df_campos = pd.DataFrame(campos_info[1:], columns=campos_info[0])
            df_campos.to_excel(writer, sheet_name='CAMPOS', index=False)
            
            # ========================================
            # HOJA 4: EJEMPLOS
            # ========================================
            ejemplos_datos = {
                'RUC': ['20123456789', '20987654321', '21212121212'],
                'Razón Social Principal': ['TRANSPORTES PUNO S.A.C.', 'LOGÍSTICA AREQUIPA E.I.R.L.', 'EMPRESA MÍNIMA S.A.C.'],
                'Dirección Fiscal': ['AV. EJERCITO 123, PUNO', 'JR. MERCADERES 456, AREQUIPA', ''],
                'Teléfono Contacto': ['051-123456 051-999888', '054-987654', ''],
                'Email Contacto': ['contacto@transportespuno.com', 'info@logisticaarequipa.com', ''],
                'Nombres Representante': ['JUAN CARLOS', 'MARIA ELENA', ''],
                'Apellidos Representante': ['MAMANI QUISPE', 'RODRIGUEZ VARGAS', ''],
                'DNI Representante': ['12345678', '87654321', ''],
                'Partida Registral': ['00123456', '00765432', ''],
                'Razón Social SUNAT': ['TRANSPORTES PUNO SOCIEDAD ANONIMA CERRADA', '', ''],
                'Razón Social Mínimo': ['TRANSPORTES PUNO', 'LOGISTICA AREQUIPA', ''],
                'Estado': ['AUTORIZADA', 'EN_TRAMITE', ''],
                'Estado SUNAT': ['ACTIVO', 'ACTIVO', ''],
                'Tipo de Servicio': ['PERSONAS', 'TURISMO', ''],
                'Observaciones': ['Empresa completa con todos los datos', 'Empresa completa con todos los datos', 'Solo datos mínimos: RUC + Razón Social']
            }
            
            df_ejemplos = pd.DataFrame(ejemplos_datos)
            df_ejemplos.to_excel(writer, sheet_name='EJEMPLOS', index=False)
            
            # Agregar comentarios explicativos en la hoja de ejemplos
            worksheet_ejemplos = writer.sheets['EJEMPLOS']
            worksheet_ejemplos['A18'] = 'EXPLICACIÓN DE EJEMPLOS:'
            worksheet_ejemplos['A19'] = 'Fila 2: Empresa completa con todos los datos opcionales'
            worksheet_ejemplos['A20'] = 'Fila 3: Empresa completa con todos los datos opcionales'
            worksheet_ejemplos['A21'] = 'Fila 4: Empresa con DATOS MÍNIMOS (solo RUC + Razón Social)'
            worksheet_ejemplos['A22'] = ''
            worksheet_ejemplos['A23'] = 'CAMPOS OBLIGATORIOS:'
            worksheet_ejemplos['A24'] = '• Solo RUC y Razón Social Principal son obligatorios'
            worksheet_ejemplos['A25'] = '• Todos los demás campos son opcionales'
            worksheet_ejemplos['A26'] = ''
            worksheet_ejemplos['A27'] = 'NORMALIZACIÓN AUTOMÁTICA:'
            worksheet_ejemplos['A28'] = '• DNI "1234567" se convierte en "01234567" (8 dígitos)'
            worksheet_ejemplos['A29'] = '• Partida "123" se convierte en "00000123" (8 dígitos mínimo)'
            worksheet_ejemplos['A30'] = '• Teléfono "051-123456 051-999888" se convierte en "051-123456, 051-999888"'
            
            # ========================================
            # FORMATEAR TODAS LAS HOJAS
            # ========================================
            
            # Formatear hoja DATOS
            worksheet_datos = writer.sheets['DATOS']
            self._formatear_hoja_datos(worksheet_datos)
            
            # Formatear hoja INSTRUCCIONES
            worksheet_instrucciones = writer.sheets['INSTRUCCIONES']
            self._formatear_hoja_instrucciones(worksheet_instrucciones)
            
            # Formatear hoja CAMPOS
            worksheet_campos = writer.sheets['CAMPOS']
            self._formatear_hoja_campos(worksheet_campos)
            
            # Formatear hoja EJEMPLOS
            worksheet_ejemplos = writer.sheets['EJEMPLOS']
            self._formatear_hoja_ejemplos(worksheet_ejemplos)
        
        buffer.seek(0)
        return buffer
    
    def generar_excel_empresas(self, empresas: List[Dict[str, Any]], columnas_visibles: Optional[List[str]] = None) -> BytesIO:
        """Generar archivo Excel con datos de empresas para exportación"""
        buffer = BytesIO()
        
        # Mapeo de columnas del frontend a campos de datos
        mapeo_columnas = {
            'ruc': 'RUC',
            'razonSocial': 'Razón Social',
            'estado': 'Estado',
            'tipoServicio': 'Tipo de Servicio',
            'direccion': 'Dirección Fiscal',
            'telefono': 'Teléfono Contacto',
            'email': 'Email Contacto',
            'representanteLegal': 'Representante Legal',
            'fechaRegistro': 'Fecha Registro',
            'rutas': 'Rutas',
            'vehiculos': 'Vehículos',
            'conductores': 'Conductores'
        }
        
        # Preparar datos para Excel
        datos_excel = []
        for empresa in empresas:
            fila_datos = {
                'RUC': empresa.get('ruc', ''),
                'Razón Social': empresa.get('razonSocial', {}).get('principal', ''),
                'Estado': self._limpiar_valor_enum(empresa.get('estado', '')),
                'Tipo de Servicio': self._limpiar_valor_enum(empresa.get('tipoServicio', '')),
                'Dirección Fiscal': empresa.get('direccionFiscal', ''),
                'Representante Legal': f"{empresa.get('representanteLegal', {}).get('nombres', '')} {empresa.get('representanteLegal', {}).get('apellidos', '')}".strip(),
                'Email Contacto': empresa.get('emailContacto', ''),
                'Teléfono Contacto': empresa.get('telefonoContacto', ''),
                'Sitio Web': empresa.get('sitioWeb', ''),
                'Fecha Registro': empresa.get('fechaRegistro', ''),
                'Resoluciones': len(empresa.get('resolucionesPrimigeniasIds', [])),
                'Vehículos': len(empresa.get('vehiculosHabilitadosIds', [])),
                'Conductores': len(empresa.get('conductoresHabilitadosIds', [])),
                'Rutas': len(empresa.get('rutasAutorizadasIds', [])),
                'Score Riesgo': empresa.get('scoreRiesgo', ''),
                'Observaciones': empresa.get('observaciones', '')
            }
            
            # Si hay columnas específicas, filtrar solo esas
            if columnas_visibles:
                fila_filtrada = {}
                for col_frontend in columnas_visibles:
                    if col_frontend in mapeo_columnas:
                        campo_excel = mapeo_columnas[col_frontend]
                        if campo_excel in fila_datos:
                            fila_filtrada[campo_excel] = fila_datos[campo_excel]
                datos_excel.append(fila_filtrada)
            else:
                datos_excel.append(fila_datos)
        
        # Crear DataFrame y escribir a Excel
        df = pd.DataFrame(datos_excel)
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Empresas', index=False)
            
            # Ajustar ancho de columnas
            worksheet = writer.sheets['Empresas']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        return buffer
    
    def generar_csv_empresas(self, empresas: List[Dict[str, Any]], columnas_visibles: Optional[List[str]] = None) -> str:
        """Generar archivo CSV con datos de empresas para exportación"""
        import csv
        from io import StringIO
        
        output = StringIO()
        
        # Mapeo de columnas del frontend a campos de datos
        mapeo_columnas = {
            'ruc': 'RUC',
            'razonSocial': 'Razón Social',
            'estado': 'Estado',
            'tipoServicio': 'Tipo de Servicio',
            'direccion': 'Dirección Fiscal',
            'telefono': 'Teléfono Contacto',
            'email': 'Email Contacto',
            'representanteLegal': 'Representante Legal',
            'fechaRegistro': 'Fecha Registro',
            'rutas': 'Rutas',
            'vehiculos': 'Vehículos',
            'conductores': 'Conductores'
        }
        
        # Definir campos según columnas visibles
        if columnas_visibles:
            fieldnames = [mapeo_columnas[col] for col in columnas_visibles if col in mapeo_columnas]
        else:
            fieldnames = [
                'RUC', 'Razón Social', 'Estado', 'Tipo de Servicio', 'Dirección Fiscal',
                'Representante Legal', 'Email Contacto', 'Teléfono Contacto', 'Sitio Web',
                'Fecha Registro', 'Resoluciones', 'Vehículos', 'Conductores', 'Rutas',
                'Score Riesgo', 'Observaciones'
            ]
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for empresa in empresas:
            fila_completa = {
                'RUC': empresa.get('ruc', ''),
                'Razón Social': empresa.get('razonSocial', {}).get('principal', ''),
                'Estado': self._limpiar_valor_enum(empresa.get('estado', '')),
                'Tipo de Servicio': self._limpiar_valor_enum(empresa.get('tipoServicio', '')),
                'Dirección Fiscal': empresa.get('direccionFiscal', ''),
                'Representante Legal': f"{empresa.get('representanteLegal', {}).get('nombres', '')} {empresa.get('representanteLegal', {}).get('apellidos', '')}".strip(),
                'Email Contacto': empresa.get('emailContacto', ''),
                'Teléfono Contacto': empresa.get('telefonoContacto', ''),
                'Sitio Web': empresa.get('sitioWeb', ''),
                'Fecha Registro': empresa.get('fechaRegistro', ''),
                'Resoluciones': len(empresa.get('resolucionesPrimigeniasIds', [])),
                'Vehículos': len(empresa.get('vehiculosHabilitadosIds', [])),
                'Conductores': len(empresa.get('conductoresHabilitadosIds', [])),
                'Rutas': len(empresa.get('rutasAutorizadasIds', [])),
                'Score Riesgo': empresa.get('scoreRiesgo', ''),
                'Observaciones': empresa.get('observaciones', '')
            }
            
            # Filtrar solo los campos que están en fieldnames
            fila_filtrada = {campo: fila_completa.get(campo, '') for campo in fieldnames}
            writer.writerow(fila_filtrada)
        
        return output.getvalue()
    
    def _formatear_hoja_datos(self, worksheet):
        """Formatear la hoja de datos"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Formatear encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 40)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _formatear_hoja_instrucciones(self, worksheet):
        """Formatear la hoja de instrucciones"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Título principal
        worksheet['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        worksheet['A1'].fill = PatternFill(start_color="D9534F", end_color="D9534F", fill_type="solid")
        worksheet['A1'].alignment = Alignment(horizontal="center")
        
        # Ajustar ancho de columna
        worksheet.column_dimensions['A'].width = 80
        
        # Formatear subtítulos
        for row in range(1, 35):
            cell = worksheet[f'A{row}']
            if cell.value and str(cell.value).endswith(':'):
                cell.font = Font(bold=True, color="366092")
    
    def _formatear_hoja_campos(self, worksheet):
        """Formatear la hoja de campos"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Formatear encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="5CB85C", end_color="5CB85C", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar ancho de columnas
        column_widths = {'A': 25, 'B': 15, 'C': 20, 'D': 35, 'E': 30}
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Colorear campos obligatorios
        obligatorio_fill = PatternFill(start_color="FCF8E3", end_color="FCF8E3", fill_type="solid")
        for row in range(2, 18):
            if worksheet[f'B{row}'].value and 'SÍ' in str(worksheet[f'B{row}'].value):
                for col in ['A', 'B', 'C', 'D', 'E']:
                    worksheet[f'{col}{row}'].fill = obligatorio_fill
    
    def _formatear_hoja_ejemplos(self, worksheet):
        """Formatear la hoja de ejemplos"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Formatear encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="F0AD4E", end_color="F0AD4E", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 35)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Formatear explicaciones
        for row in range(18, 22):
            worksheet[f'A{row}'].font = Font(bold=True, color="366092")
    
    async def validar_archivo_excel(self, archivo_excel: BytesIO) -> Dict[str, Any]:
        """Validar archivo Excel de empresas usando datos reales de la base de datos"""
        try:
            # Intentar leer la hoja "DATOS" primero, si no existe, leer la primera hoja
            try:
                df = pd.read_excel(archivo_excel, sheet_name='DATOS')
            except:
                # Si no existe la hoja DATOS, leer la primera hoja disponible
                df = pd.read_excel(archivo_excel)
            
            resultados = {
                'total_filas': len(df),
                'validos': 0,
                'invalidos': 0,
                'con_advertencias': 0,
                'errores': [],
                'advertencias': [],
                'empresas_validas': []
            }
            
            for index, row in df.iterrows():
                fila_num = index + 2  # +2 porque Excel empieza en 1 y tiene header
                errores_fila = []
                advertencias_fila = []
                
                # Validar fila
                errores_fila, advertencias_fila = await self._validar_fila_empresa(row, fila_num)
                
                if errores_fila:
                    resultados['invalidos'] += 1
                    ruc = str(row.get('RUC', 'N/A')).strip()
                    resultados['errores'].append({
                        'fila': fila_num,
                        'ruc': ruc,
                        'errores': errores_fila
                    })
                else:
                    if advertencias_fila:
                        resultados['con_advertencias'] += 1
                        ruc = str(row.get('RUC', 'N/A')).strip()
                        resultados['advertencias'].append({
                            'fila': fila_num,
                            'ruc': ruc,
                            'advertencias': advertencias_fila
                        })
                    
                    resultados['validos'] += 1
                    # Convertir fila a datos de empresa
                    try:
                        empresa_data = self._convertir_fila_a_empresa_update(row)
                        resultados['empresas_validas'].append(empresa_data)
                    except Exception as e:
                        resultados['validos'] -= 1
                        resultados['invalidos'] += 1
                        ruc = str(row.get('RUC', 'N/A')).strip()
                        resultados['errores'].append({
                            'fila': fila_num,
                            'ruc': ruc,
                            'errores': [f"Error al procesar empresa: {str(e)}"]
                        })
            
            return resultados
            
        except Exception as e:
            return {
                'error': f"Error al procesar archivo Excel: {str(e)}",
                'total_filas': 0,
                'validos': 0,
                'invalidos': 0,
                'con_advertencias': 0,
                'errores': [],
                'advertencias': [],
                'empresas_validas': []
            }
    
    async def _validar_fila_empresa(self, row: pd.Series, fila_num: int) -> Tuple[List[str], List[str]]:
        """Validar una fila de empresa contra la base de datos real"""
        errores = []
        advertencias = []
        
        def limpiar_valor(valor):
            """Limpiar y convertir valores de pandas"""
            if pd.isna(valor):
                return None
            valor_str = str(valor).strip()
            if valor_str == '' or valor_str.lower() == 'nan':
                return None
            # Si es un número float que termina en .0, convertir a entero
            if valor_str.endswith('.0'):
                try:
                    return str(int(float(valor_str)))
                except:
                    return valor_str
            return valor_str
        
        # Validar RUC (OBLIGATORIO)
        ruc = str(row.get('RUC', '')).strip() if pd.notna(row.get('RUC')) else ''
        if not ruc:
            errores.append("RUC es requerido")
        elif not self._validar_formato_ruc(ruc):
            errores.append(f"RUC debe tener exactamente 11 dígitos: {ruc}")
        else:
            # Verificar si ya existe en la base de datos REAL
            if await self._existe_empresa_con_ruc(ruc):
                advertencias.append(f"Empresa con RUC {ruc} ya existe - se actualizará con los nuevos datos")
        
        # Validar razón social principal (OBLIGATORIO)
        razon_social = str(row.get('Razón Social Principal', '')).strip() if pd.notna(row.get('Razón Social Principal')) else ''
        if not razon_social:
            errores.append("Razón Social Principal es requerida")
        elif len(razon_social) < 3:
            errores.append("Razón Social Principal debe tener al menos 3 caracteres")
        
        # Validar dirección fiscal (OPCIONAL)
        direccion = str(row.get('Dirección Fiscal', '')).strip() if pd.notna(row.get('Dirección Fiscal')) else ''
        if direccion and len(direccion) < 10:
            errores.append("Dirección Fiscal debe tener al menos 10 caracteres")
        
        # Validar estado (OPCIONAL)
        estado_raw = row.get('Estado', '')
        if pd.isna(estado_raw) or str(estado_raw).strip() == '':
            estado = 'AUTORIZADA'  # Estado por defecto
        else:
            estado = str(estado_raw).strip().upper()
        
        if estado and estado not in [e.value for e in EstadoEmpresa]:
            errores.append(f"Estado inválido: {estado}. Valores válidos: {', '.join([e.value for e in EstadoEmpresa])}")
        
        # Validar DNI representante (OPCIONAL)
        dni_rep = limpiar_valor(row.get('DNI Representante', ''))
        if dni_rep and not self._validar_formato_dni(dni_rep):
            errores.append(f"DNI debe ser numérico y tener máximo 8 dígitos: {dni_rep}")
        
        # Validar nombres representante (OPCIONAL)
        nombres_rep = limpiar_valor(row.get('Nombres Representante', ''))
        if nombres_rep and len(nombres_rep) < 2:
            errores.append("Nombres del Representante deben tener al menos 2 caracteres")
        
        # Validar apellidos representante (OPCIONAL)
        apellidos_rep = limpiar_valor(row.get('Apellidos Representante', ''))
        if apellidos_rep and len(apellidos_rep) < 2:
            errores.append("Apellidos del Representante deben tener al menos 2 caracteres")
        
        # Validar teléfono contacto (opcional pero formato válido si se proporciona)
        telefono_contacto = limpiar_valor(row.get('Teléfono Contacto', ''))
        if telefono_contacto and not self._validar_formato_telefono(telefono_contacto):
            errores.append(f"Formato de teléfono de contacto inválido: {telefono_contacto}")
        
        # Validar email contacto (opcional pero formato válido si se proporciona)
        email_contacto = limpiar_valor(row.get('Email Contacto', ''))
        if email_contacto and not self._validar_formato_email(email_contacto):
            errores.append(f"Formato de email de contacto inválido: {email_contacto}")
        
        # Validar partida registral (opcional)
        partida_registral = limpiar_valor(row.get('Partida Registral', ''))
        if partida_registral and not self._validar_formato_partida_registral(partida_registral):
            errores.append(f"Partida Registral debe ser numérica y tener entre 1 y 9 dígitos: {partida_registral}")
        
        # Validar razón social SUNAT (opcional)
        razon_social_sunat = limpiar_valor(row.get('Razón Social SUNAT', ''))
        if razon_social_sunat and len(razon_social_sunat) < 3:
            errores.append("Razón Social SUNAT debe tener al menos 3 caracteres")
        
        # Validar razón social mínimo (opcional)
        razon_social_minimo = limpiar_valor(row.get('Razón Social Mínimo', ''))
        if razon_social_minimo and len(razon_social_minimo) < 2:
            errores.append("Razón Social Mínimo debe tener al menos 2 caracteres")
        
        # Validar estado SUNAT (opcional)
        estado_sunat = limpiar_valor(row.get('Estado SUNAT', ''))
        if estado_sunat and len(estado_sunat) < 2:
            errores.append("Estado SUNAT debe tener al menos 2 caracteres")
        
        # Validar tipo de servicio (opcional)
        tipo_servicio = limpiar_valor(row.get('Tipo de Servicio', ''))
        if tipo_servicio:
            # Obtener tipos válidos desde configuraciones
            try:
                config_service = await self._get_configuracion_service()
                if config_service:
                    tipos_validos = await config_service.get_tipos_servicio_codigos()
                    if tipo_servicio.upper() not in tipos_validos:
                        errores.append(f"Tipo de Servicio inválido: {tipo_servicio}. Valores válidos: {', '.join(tipos_validos)}")
                else:
                    # Si no hay conexión a BD, usar tipos básicos para validación
                    tipos_basicos = ['PERSONAS', 'TURISMO', 'MERCANCIAS', 'CARGA']
                    if tipo_servicio.upper() not in tipos_basicos:
                        advertencias.append(f"Tipo de Servicio no validado contra BD: {tipo_servicio}")
            except Exception as e:
                print(f"Error validando tipo de servicio: {e}")
                # En caso de error, solo advertir
                advertencias.append(f"No se pudo validar Tipo de Servicio: {tipo_servicio}")
        
        return errores, advertencias
    
    def _validar_formato_ruc(self, ruc: str) -> bool:
        """Validar formato de RUC: 11 dígitos"""
        return ruc.isdigit() and len(ruc) == 11
    
    def _validar_formato_dni(self, dni: str) -> bool:
        """Validar formato de DNI: debe ser numérico y tener máximo 8 dígitos"""
        return dni.isdigit() and len(dni) <= 8
    
    def _validar_formato_email(self, email: str) -> bool:
        """Validar formato de email"""
        patron = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(patron, email))
    
    def _validar_formato_telefono(self, telefono: str) -> bool:
        """Validar formato de teléfono: números, espacios, guiones y paréntesis. Acepta múltiples teléfonos separados por espacios"""
        if not telefono:
            return True
        
        # Normalizar el teléfono para validación
        telefono_normalizado = self._normalizar_telefono(telefono)
        
        # Si contiene comas, validar cada número por separado
        if ',' in telefono_normalizado:
            numeros = [num.strip() for num in telefono_normalizado.split(',')]
            for numero in numeros:
                if numero:  # Solo validar números no vacíos
                    patron = r'^[\d\s\-\(\)\+]{7,15}$'
                    if not re.match(patron, numero):
                        return False
            return True
        else:
            # Validar como un solo número
            patron = r'^[\d\s\-\(\)\+]{7,15}$'
            return bool(re.match(patron, telefono_normalizado))
    
    def _procesar_nombres_apellidos(self, nombres_completos: str, apellidos: str = None) -> tuple:
        """
        Procesar nombres y apellidos, manejando casos donde todo viene en el campo nombres
        
        Args:
            nombres_completos: Campo nombres que puede contener nombre completo
            apellidos: Campo apellidos (puede estar vacío)
            
        Returns:
            tuple: (nombres_procesados, apellidos_procesados)
        """
        if not nombres_completos:
            return None, None
        
        nombres_completos = nombres_completos.strip()
        
        # Si hay apellidos explícitos, usarlos
        if apellidos and apellidos.strip():
            return nombres_completos, apellidos.strip()
        
        # Si no hay apellidos, intentar dividir el nombre completo
        partes = nombres_completos.split()
        
        if len(partes) >= 2:
            # Asumir que las primeras palabras son nombres y las últimas apellidos
            if len(partes) == 2:
                # Caso simple: "JUAN PEREZ" -> nombres="JUAN", apellidos="PEREZ"
                return partes[0], partes[1]
            elif len(partes) == 3:
                # "JUAN CARLOS PEREZ" -> nombres="JUAN CARLOS", apellidos="PEREZ"
                return " ".join(partes[:-1]), partes[-1]
            elif len(partes) >= 4:
                # "JUAN CARLOS PEREZ GARCIA" -> nombres="JUAN CARLOS", apellidos="PEREZ GARCIA"
                mitad = len(partes) // 2
                return " ".join(partes[:mitad]), " ".join(partes[mitad:])
        
        # Si solo hay una palabra, usar como nombres y generar apellido genérico
        return nombres_completos, "SIN APELLIDO"
    
    def _normalizar_telefono(self, telefono: str) -> str:
        """Normalizar teléfono: convertir espacios a comas para múltiples números"""
        if not telefono:
            return telefono
        
        # Limpiar el teléfono
        telefono_limpio = telefono.strip()
        
        # Si contiene espacios que separan números completos, convertir a comas
        # Patrón mejorado para detectar múltiples números separados por espacios
        import re
        
        # Buscar patrones de números telefónicos completos separados por espacios
        # Un número telefónico puede incluir: dígitos, guiones, paréntesis, signos +
        # Pero debe tener al menos 7 dígitos consecutivos o separados por guiones/paréntesis
        
        # Dividir por espacios múltiples (2 o más) que separan números completos
        # O por espacios simples que están entre números que parecen completos
        
        # Estrategia más simple: dividir por espacios y verificar si cada parte parece un número
        partes = re.split(r'\s+', telefono_limpio)
        
        if len(partes) > 1:
            # Verificar si cada parte parece un número telefónico válido
            numeros_validos = []
            for parte in partes:
                parte = parte.strip()
                if parte and len(parte) >= 7:  # Mínimo 7 caracteres para ser un número
                    # Verificar que tenga suficientes dígitos
                    digitos = re.findall(r'\d', parte)
                    if len(digitos) >= 6:  # Al menos 6 dígitos
                        numeros_validos.append(parte)
                    else:
                        # Si no es un número válido, mantener el original
                        return telefono_limpio
            
            # Si todos son números válidos, unir con comas
            if len(numeros_validos) == len(partes):
                return ', '.join(numeros_validos)
        
        # Si no se detectaron múltiples números, devolver original
        return telefono_limpio
        
        # Buscar patrones de números telefónicos completos separados por espacios
        # Un número telefónico puede incluir: dígitos, guiones, paréntesis, signos +
        # Pero debe tener al menos 7 dígitos consecutivos o separados por guiones/paréntesis
        
        # Patrón para identificar un número telefónico completo
        patron_numero = r'[\+]?[\d\(\)\-\s]{7,15}'
        
        # Dividir por espacios múltiples (2 o más) que separan números completos
        # O por espacios simples que están entre números que parecen completos
        
        # Estrategia más simple: dividir por espacios y verificar si cada parte parece un número
        partes = re.split(r'\s+', telefono_limpio)
        
        if len(partes) > 1:
            # Verificar si cada parte parece un número telefónico válido
            numeros_validos = []
            for parte in partes:
                parte = parte.strip()
                if parte and len(parte) >= 7:  # Mínimo 7 caracteres para ser un número
                    # Verificar que tenga suficientes dígitos
                    digitos = re.findall(r'\d', parte)
                    if len(digitos) >= 6:  # Al menos 6 dígitos
                        numeros_validos.append(parte)
                    else:
                        # Si no es un número válido, mantener el original
                        return telefono_limpio
                elif parte:
                    # Si hay una parte muy corta, mantener el original
                    return telefono_limpio
            
            # Si todas las partes son números válidos, unir con comas
            if len(numeros_validos) > 1:
                return ', '.join(numeros_validos)
        
        return telefono_limpio
    
    def _validar_formato_partida_registral(self, partida: str) -> bool:
        """Validar formato de partida registral: debe ser numérico y tener entre 1 y 9 dígitos"""
        return partida.isdigit() and 1 <= len(partida) <= 9
    
    def _normalizar_dni(self, dni: str) -> str:
        """Normalizar DNI completando con ceros a la izquierda hasta 8 dígitos"""
        if not dni or not dni.isdigit():
            return dni
        return dni.zfill(8)
    
    def _normalizar_partida_registral(self, partida: str) -> str:
        """Normalizar partida registral completando con ceros a la izquierda hasta 8 dígitos mínimo"""
        if not partida or not partida.isdigit():
            return partida
        # Si tiene menos de 8 dígitos, completar con ceros hasta 8
        if len(partida) < 8:
            return partida.zfill(8)
        # Si tiene 8 o 9 dígitos, mantener como está
        return partida
    
    async def _existe_empresa_con_ruc(self, ruc: str) -> bool:
        """Verificar si existe empresa con el RUC dado en la base de datos REAL"""
        try:
            empresa_service = await self._get_empresa_service()
            if empresa_service is None:
                # Si no hay conexión a BD, asumir que no existe (para pruebas)
                return False
            empresa = await empresa_service.get_empresa_by_ruc(ruc)
            return empresa is not None
        except Exception as e:
            print(f"Error verificando RUC empresa: {e}")
            # En caso de error, asumir que no existe para permitir la validación
            return False
    
    def _convertir_fila_a_empresa_update(self, row: pd.Series) -> dict:
        """Convertir fila de Excel a diccionario para actualización (solo campos no vacíos)"""
        
        def limpiar_valor(valor):
            """Limpiar y convertir valores de pandas"""
            if pd.isna(valor):
                return None
            valor_str = str(valor).strip()
            if valor_str == '' or valor_str.lower() == 'nan':
                return None
            # Si es un número float que termina en .0, convertir a entero
            if valor_str.endswith('.0'):
                try:
                    return str(int(float(valor_str)))
                except:
                    return valor_str
            return valor_str
        
        # Datos básicos
        ruc = limpiar_valor(row.get('RUC', ''))
        
        # Razón social (solo si se proporciona)
        razon_social_principal = limpiar_valor(row.get('Razón Social Principal', ''))
        razon_social_sunat = limpiar_valor(row.get('Razón Social SUNAT', ''))
        razon_social_minimo = limpiar_valor(row.get('Razón Social Mínimo', ''))
        
        update_data = {'ruc': ruc}
        
        if razon_social_principal:  # Solo crear si hay razón social principal
            razon_social = RazonSocial(
                principal=razon_social_principal,
                sunat=razon_social_sunat,
                minimo=razon_social_minimo
            )
            update_data['razonSocial'] = razon_social
        
        # Dirección fiscal (solo si se proporciona)
        direccion_fiscal = limpiar_valor(row.get('Dirección Fiscal', ''))
        if direccion_fiscal:
            update_data['direccionFiscal'] = direccion_fiscal
        
        # Contacto empresa (solo si se proporciona)
        telefono_contacto = limpiar_valor(row.get('Teléfono Contacto', ''))
        if telefono_contacto:
            # Normalizar teléfono: convertir espacios a comas para múltiples números
            telefono_normalizado = self._normalizar_telefono(telefono_contacto)
            update_data['telefonoContacto'] = telefono_normalizado
            
        email_contacto = limpiar_valor(row.get('Email Contacto', ''))
        if email_contacto:
            update_data['emailContacto'] = email_contacto
        
        # Representante legal (solo requiere DNI, nombres y apellidos opcionales)
        nombres_rep = limpiar_valor(row.get('Nombres Representante', ''))
        apellidos_rep = limpiar_valor(row.get('Apellidos Representante', ''))
        dni_rep = limpiar_valor(row.get('DNI Representante', ''))
        
        if dni_rep:  # Solo requiere DNI
            # Normalizar DNI completando con ceros a la izquierda
            dni_normalizado = self._normalizar_dni(dni_rep)
            
            # Usar valores por defecto si no se proporcionan nombres/apellidos
            nombres_final = nombres_rep if nombres_rep else "POR ACTUALIZAR"
            apellidos_final = apellidos_rep if apellidos_rep else "DESDE API EXTERNA"
            
            representante_legal = RepresentanteLegal(
                dni=dni_normalizado,
                nombres=nombres_final,
                apellidos=apellidos_final,
                email=None,  # No hay email representante en el nuevo formato
                telefono=None,  # No hay teléfono representante en el nuevo formato
                direccion=None  # No hay dirección representante en el nuevo formato
            )
            update_data['representanteLegal'] = representante_legal
        
        # Campos adicionales nuevos
        partida_registral = limpiar_valor(row.get('Partida Registral', ''))
        if partida_registral:
            # Normalizar partida registral completando con ceros a la izquierda
            partida_normalizada = self._normalizar_partida_registral(partida_registral)
            update_data['partidaRegistral'] = partida_normalizada
        
        estado_sunat = limpiar_valor(row.get('Estado SUNAT', ''))
        if estado_sunat:
            update_data['estadoSunat'] = estado_sunat
        
        # Observaciones (solo si se proporciona)
        observaciones = limpiar_valor(row.get('Observaciones', ''))
        if observaciones:
            update_data['observaciones'] = observaciones
        
        # Tipo de servicio (solo si se proporciona)
        tipo_servicio = limpiar_valor(row.get('Tipo de Servicio', ''))
        if tipo_servicio:
            update_data['tipoServicio'] = tipo_servicio.upper()
        
        return update_data
        
    
    def _convertir_fila_a_empresa_create(self, row: pd.Series) -> EmpresaCreate:
        """Convertir fila de Excel a modelo EmpresaCreate (solo campos no vacíos)"""
        
        def limpiar_valor(valor):
            """Limpiar y convertir valores de pandas"""
            if pd.isna(valor):
                return None
            valor_str = str(valor).strip()
            if valor_str == '' or valor_str.lower() == 'nan':
                return None
            # Si es un número float que termina en .0, convertir a entero
            if valor_str.endswith('.0'):
                try:
                    return str(int(float(valor_str)))
                except:
                    return valor_str
            return valor_str
        
        # Datos básicos
        ruc = limpiar_valor(row.get('RUC', ''))
        
        # Razón social (solo si se proporciona)
        razon_social_principal = limpiar_valor(row.get('Razón Social Principal', ''))
        razon_social_sunat = limpiar_valor(row.get('Razón Social SUNAT', ''))
        razon_social_minimo = limpiar_valor(row.get('Razón Social Mínimo', ''))
        
        razon_social = None
        if razon_social_principal:  # Solo crear si hay razón social principal
            razon_social = RazonSocial(
                principal=razon_social_principal,
                sunat=razon_social_sunat,
                minimo=razon_social_minimo
            )
        
        # Dirección fiscal (solo si se proporciona)
        direccion_fiscal = limpiar_valor(row.get('Dirección Fiscal', ''))
        
        # Contacto empresa (solo si se proporciona)
        telefono_contacto = limpiar_valor(row.get('Teléfono Contacto', ''))
        email_contacto = limpiar_valor(row.get('Email Contacto', ''))
        
        # Representante legal (solo requiere DNI, nombres y apellidos opcionales)
        nombres_rep = limpiar_valor(row.get('Nombres Representante', ''))
        apellidos_rep = limpiar_valor(row.get('Apellidos Representante', ''))
        dni_rep = limpiar_valor(row.get('DNI Representante', ''))
        
        representante_legal = None
        if dni_rep:  # Solo requiere DNI, nombres y apellidos opcionales
            # Normalizar DNI completando con ceros a la izquierda
            dni_normalizado = self._normalizar_dni(dni_rep)
            
            # Usar valores por defecto si no se proporcionan nombres/apellidos
            nombres_final = nombres_rep if nombres_rep else "POR ACTUALIZAR"
            apellidos_final = apellidos_rep if apellidos_rep else "DESDE API EXTERNA"
            
            representante_legal = RepresentanteLegal(
                dni=dni_normalizado,
                nombres=nombres_final,
                apellidos=apellidos_final,
                email=None,  # No hay email representante en el nuevo formato
                telefono=None,  # No hay teléfono representante en el nuevo formato
                direccion=None  # No hay dirección representante en el nuevo formato
            )
        
        # Observaciones (solo si se proporciona)
        observaciones = limpiar_valor(row.get('Observaciones', ''))
        
        # Crear objeto con solo los campos que tienen datos
        empresa_data = {
            'ruc': ruc
        }
        
        if razon_social:
            empresa_data['razonSocial'] = razon_social
        if direccion_fiscal:
            empresa_data['direccionFiscal'] = direccion_fiscal
        if representante_legal:
            empresa_data['representanteLegal'] = representante_legal
        if email_contacto:
            empresa_data['emailContacto'] = email_contacto
        if telefono_contacto:
            # Normalizar teléfono: convertir espacios a comas para múltiples números
            telefono_normalizado = self._normalizar_telefono(telefono_contacto)
            empresa_data['telefonoContacto'] = telefono_normalizado
        if observaciones:
            empresa_data['observaciones'] = observaciones
        
        # Tipo de servicio (solo si se proporciona)
        tipo_servicio = limpiar_valor(row.get('Tipo de Servicio', ''))
        if tipo_servicio:
            empresa_data['tipoServicio'] = tipo_servicio.upper()
        
        return EmpresaCreate(**empresa_data)
        ruc = limpiar_valor(row.get('RUC', ''))
        
        # Razón social (solo si se proporciona)
        razon_social_principal = limpiar_valor(row.get('Razón Social Principal', ''))
        razon_social_sunat = limpiar_valor(row.get('Razón Social SUNAT', ''))
        razon_social_minimo = limpiar_valor(row.get('Razón Social Mínimo', ''))
        
        razon_social = None
        if razon_social_principal:  # Solo crear si hay razón social principal
            razon_social = RazonSocial(
                principal=razon_social_principal,
                sunat=razon_social_sunat,
                minimo=razon_social_minimo
            )
        
        # Dirección fiscal (solo si se proporciona)
        direccion_fiscal = limpiar_valor(row.get('Dirección Fiscal', ''))
        
        # Representante legal (solo si se proporcionan datos básicos)
        dni_rep = limpiar_valor(row.get('DNI Representante', ''))
        nombres_rep = limpiar_valor(row.get('Nombres Representante', ''))
        apellidos_rep = limpiar_valor(row.get('Apellidos Representante', ''))
        email_rep = limpiar_valor(row.get('Email Representante', ''))
        telefono_rep = limpiar_valor(row.get('Teléfono Representante', ''))
        direccion_rep = limpiar_valor(row.get('Dirección Representante', ''))
        
        representante_legal = None
        if dni_rep and nombres_rep and apellidos_rep:  # Solo crear si hay datos básicos
            representante_legal = RepresentanteLegal(
                dni=dni_rep,
                nombres=nombres_rep,
                apellidos=apellidos_rep,
                email=email_rep,
                telefono=telefono_rep,
                direccion=direccion_rep
            )
        
        # Contacto empresa (solo si se proporciona)
        email_contacto = limpiar_valor(row.get('Email Contacto', ''))
        telefono_contacto = limpiar_valor(row.get('Teléfono Contacto', ''))
        sitio_web = limpiar_valor(row.get('Sitio Web', ''))
        
        # Observaciones (solo si se proporciona)
        observaciones = limpiar_valor(row.get('Observaciones', ''))
        
        # Crear objeto con solo los campos que tienen datos
        empresa_data = {
            'ruc': ruc
        }
        
        if razon_social:
            empresa_data['razonSocial'] = razon_social
        if direccion_fiscal:
            empresa_data['direccionFiscal'] = direccion_fiscal
        if representante_legal:
            empresa_data['representanteLegal'] = representante_legal
        if email_contacto:
            empresa_data['emailContacto'] = email_contacto
        if telefono_contacto:
            # Normalizar teléfono: convertir espacios a comas para múltiples números
            telefono_normalizado = self._normalizar_telefono(telefono_contacto)
            empresa_data['telefonoContacto'] = telefono_normalizado
        if sitio_web:
            empresa_data['sitioWeb'] = sitio_web
        if observaciones:
            empresa_data['observaciones'] = observaciones
        
        return EmpresaCreate(**empresa_data)
    
    async def procesar_carga_masiva(self, archivo_excel: BytesIO) -> Dict[str, Any]:
        """Procesar carga masiva de empresas desde Excel - CREAR O ACTUALIZAR EN BASE DE DATOS REAL"""
        
        # Primero validar el archivo
        resultado_validacion = await self.validar_archivo_excel(archivo_excel)
        
        if 'error' in resultado_validacion:
            return resultado_validacion
        
        # Si hay empresas válidas, procesarlas en la base de datos REAL
        empresas_creadas = []
        empresas_actualizadas = []
        errores_creacion = []
        
        empresa_service = await self._get_empresa_service()
        
        for empresa_data in resultado_validacion['empresas_validas']:
            try:
                # Verificar si la empresa ya existe
                ruc = empresa_data['ruc']
                empresa_existente = await empresa_service.get_empresa_by_ruc(ruc)
                
                if empresa_existente:
                    # ACTUALIZAR empresa existente
                    empresa_actualizada = await self._actualizar_empresa_existente_dict(
                        empresa_existente, empresa_data, empresa_service
                    )
                    
                    empresas_actualizadas.append({
                        'ruc': empresa_actualizada.ruc,
                        'razon_social': empresa_actualizada.razonSocial.principal,
                        'estado': empresa_actualizada.estado.value,
                        'accion': 'ACTUALIZADA'
                    })
                else:
                    # CREAR nueva empresa - convertir dict a EmpresaCreate (SIN validaciones externas)
                    empresa_create = self._dict_to_empresa_create(empresa_data)
                    usuario_id = "CARGA_MASIVA"
                    empresa_creada = await empresa_service.create_empresa_carga_masiva(empresa_create, usuario_id)
                    
                    empresas_creadas.append({
                        'ruc': empresa_creada.ruc,
                        'razon_social': empresa_creada.razonSocial.principal,
                        'estado': empresa_creada.estado.value,
                        'accion': 'CREADA'
                    })
                
            except Exception as e:
                errores_creacion.append({
                    'ruc': empresa_data.get('ruc', 'N/A'),
                    'error': str(e)
                })
        
        return {
            **resultado_validacion,
            'empresas_creadas': empresas_creadas,
            'empresas_actualizadas': empresas_actualizadas,
            'errores_creacion': errores_creacion,
            'total_creadas': len(empresas_creadas),
            'total_actualizadas': len(empresas_actualizadas),
            'total_procesadas': len(empresas_creadas) + len(empresas_actualizadas)
        }
    
    async def _actualizar_empresa_existente(self, empresa_existente, empresa_data, empresa_service):
        """Actualizar empresa existente manteniendo campos vacíos del Excel"""
        from app.models.empresa import EmpresaUpdate
        
        # Crear objeto de actualización solo con campos que tienen datos
        update_data = {}
        
        # Solo actualizar campos que vienen con datos en el Excel
        if hasattr(empresa_data, 'razonSocial') and empresa_data.razonSocial:
            update_data['razonSocial'] = empresa_data.razonSocial
            
        if hasattr(empresa_data, 'direccionFiscal') and empresa_data.direccionFiscal:
            update_data['direccionFiscal'] = empresa_data.direccionFiscal
            
        if hasattr(empresa_data, 'representanteLegal') and empresa_data.representanteLegal:
            update_data['representanteLegal'] = empresa_data.representanteLegal
            
        if hasattr(empresa_data, 'emailContacto') and empresa_data.emailContacto:
            update_data['emailContacto'] = empresa_data.emailContacto
            
        if hasattr(empresa_data, 'telefonoContacto') and empresa_data.telefonoContacto:
            update_data['telefonoContacto'] = empresa_data.telefonoContacto
            
        if hasattr(empresa_data, 'sitioWeb') and empresa_data.sitioWeb:
            update_data['sitioWeb'] = empresa_data.sitioWeb
            
        if hasattr(empresa_data, 'observaciones') and empresa_data.observaciones:
            update_data['observaciones'] = empresa_data.observaciones
        
        # Si no hay datos para actualizar, devolver la empresa existente
        if not update_data:
            return empresa_existente
        
        # Crear objeto EmpresaUpdate
        empresa_update = EmpresaUpdate(**update_data)
        
        # Actualizar en la base de datos
        usuario_id = "CARGA_MASIVA_UPDATE"
        empresa_actualizada = await empresa_service.update_empresa(empresa_existente.id, empresa_update, usuario_id)
        
        return empresa_actualizada
    
    def _dict_to_empresa_create(self, empresa_dict: dict) -> EmpresaCreate:
        """Convertir diccionario a EmpresaCreate para nuevas empresas"""
        # Solo RUC y Razón Social Principal son obligatorios
        if 'ruc' not in empresa_dict or not empresa_dict['ruc']:
            raise ValueError("RUC es requerido para nuevas empresas")
        if 'razonSocial' not in empresa_dict or not empresa_dict['razonSocial']:
            raise ValueError("Razón Social es requerida para nuevas empresas")
        
        # Crear datos mínimos para nueva empresa
        empresa_data = {
            'ruc': empresa_dict['ruc'],
            'razonSocial': empresa_dict['razonSocial']
        }
        
        # Agregar campos opcionales si están presentes
        if 'direccionFiscal' in empresa_dict and empresa_dict['direccionFiscal']:
            empresa_data['direccionFiscal'] = empresa_dict['direccionFiscal']
        else:
            # Dirección por defecto si no se proporciona
            empresa_data['direccionFiscal'] = "POR ACTUALIZAR"
        
        if 'representanteLegal' in empresa_dict and empresa_dict['representanteLegal']:
            empresa_data['representanteLegal'] = empresa_dict['representanteLegal']
        else:
            # Representante por defecto si no se proporciona
            from app.models.empresa import RepresentanteLegal
            empresa_data['representanteLegal'] = RepresentanteLegal(
                dni="00000000",
                nombres="POR ACTUALIZAR",
                apellidos="DESDE API EXTERNA"
            )
        
        # Tipo de servicio por defecto
        if 'tipoServicio' in empresa_dict and empresa_dict['tipoServicio']:
            empresa_data['tipoServicio'] = empresa_dict['tipoServicio']
        else:
            from app.models.empresa import TipoServicio
            empresa_data['tipoServicio'] = TipoServicio.PERSONAS
        
        # Agregar otros campos opcionales si están presentes
        optional_fields = ['emailContacto', 'telefonoContacto', 'sitioWeb', 'observaciones']
        for field in optional_fields:
            if field in empresa_dict and empresa_dict[field]:
                empresa_data[field] = empresa_dict[field]
        
        return EmpresaCreate(**empresa_data)
    
    async def _actualizar_empresa_existente_dict(self, empresa_existente, empresa_dict, empresa_service):
        """Actualizar empresa existente usando diccionario de datos"""
        from app.models.empresa import EmpresaUpdate
        
        # Crear objeto de actualización solo con campos que tienen datos
        update_data = {}
        
        # Solo actualizar campos que vienen con datos en el diccionario
        for key, value in empresa_dict.items():
            if key != 'ruc' and value is not None:  # No actualizar RUC
                update_data[key] = value
        
        # Si no hay datos para actualizar, devolver la empresa existente
        if not update_data:
            return empresa_existente
        
        # Crear objeto EmpresaUpdate
        empresa_update = EmpresaUpdate(**update_data)
        
        # Actualizar en la base de datos
        usuario_id = "CARGA_MASIVA_UPDATE"
        empresa_actualizada = await empresa_service.update_empresa(empresa_existente.id, empresa_update, usuario_id)
        
        return empresa_actualizada