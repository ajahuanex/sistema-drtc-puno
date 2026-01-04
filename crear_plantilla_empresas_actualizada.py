#!/usr/bin/env python3
"""
Script para crear la plantilla Excel actualizada de carga masiva de empresas
con las nuevas indicaciones para múltiples teléfonos y validaciones flexibles.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def crear_plantilla_empresas_actualizada():
    """Crea la plantilla Excel actualizada con las nuevas indicaciones."""
    
    # Crear workbook
    wb = Workbook()
    
    # Hoja de instrucciones
    ws_instrucciones = wb.active
    ws_instrucciones.title = "INSTRUCCIONES"
    
    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    titulo_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    
    subtitulo_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    subtitulo_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    
    texto_font = Font(name='Arial', size=10)
    ejemplo_font = Font(name='Arial', size=9, italic=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título principal
    ws_instrucciones['A1'] = 'PLANTILLA DE CARGA MASIVA DE EMPRESAS - ACTUALIZADA'
    ws_instrucciones['A1'].font = titulo_font
    ws_instrucciones['A1'].fill = titulo_fill
    ws_instrucciones['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_instrucciones.merge_cells('A1:H1')
    
    # Instrucciones principales
    row = 3
    instrucciones = [
        ("CAMBIOS IMPORTANTES:", subtitulo_font, subtitulo_fill),
        ("", None, None),
        ("1. MÚLTIPLES TELÉFONOS:", Font(name='Arial', size=10, bold=True), None),
        ("   • Puede ingresar varios números separados por ESPACIOS", texto_font, None),
        ("   • Ejemplo: 987654321 123456789 555666777", ejemplo_font, None),
        ("   • El sistema convierte automáticamente espacios a comas", texto_font, None),
        ("", None, None),
        ("2. REPRESENTANTE LEGAL:", Font(name='Arial', size=10, bold=True), None),
        ("   • DNI: OBLIGATORIO (8 dígitos exactos)", texto_font, None),
        ("   • Nombres: OPCIONAL (se completa automáticamente si está vacío)", texto_font, None),
        ("   • Apellidos: OPCIONAL (se completa automáticamente si está vacío)", texto_font, None),
        ("", None, None),
        ("3. EMAIL INSTITUCIONAL:", Font(name='Arial', size=10, bold=True), None),
        ("   • Debe usar el dominio: @transportespuno.gob.pe", texto_font, None),
        ("   • Ejemplo: empresa@transportespuno.gob.pe", ejemplo_font, None),
        ("", None, None),
        ("4. VALIDACIONES AUTOMÁTICAS:", Font(name='Arial', size=10, bold=True), None),
        ("   • RUC: Exactamente 11 dígitos", texto_font, None),
        ("   • DNI: Exactamente 8 dígitos", texto_font, None),
        ("   • Teléfonos: Se normalizan automáticamente", texto_font, None),
        ("   • Campos opcionales: Se completan con 'Por actualizar'", texto_font, None),
    ]
    
    for instruccion, font, fill in instrucciones:
        ws_instrucciones[f'A{row}'] = instruccion
        if font:
            ws_instrucciones[f'A{row}'].font = font
        if fill:
            ws_instrucciones[f'A{row}'].fill = fill
            ws_instrucciones.merge_cells(f'A{row}:H{row}')
        row += 1
    
    # Ajustar ancho de columnas
    ws_instrucciones.column_dimensions['A'].width = 80
    
    # Hoja de datos de ejemplo
    ws_datos = wb.create_sheet("DATOS_EJEMPLO")
    
    # Encabezados
    headers = [
        'razon_social',
        'ruc', 
        'direccion',
        'telefono',
        'email',
        'representante_dni',
        'representante_nombres',
        'representante_apellidos'
    ]
    
    # Datos de ejemplo
    datos_ejemplo = [
        [
            'Transportes El Sol SAC',
            '20123456789',
            'Av. Principal 123, Puno',
            '987654321 123456789',
            'elsol@transportespuno.gob.pe',
            '12345678',
            'Juan Carlos',
            'Pérez López'
        ],
        [
            'Empresa Luna EIRL',
            '20987654321', 
            'Jr. Comercio 456, Juliaca',
            '555666777',
            'luna@transportespuno.gob.pe',
            '87654321',
            '',  # Nombres vacío - se completará automáticamente
            ''   # Apellidos vacío - se completará automáticamente
        ],
        [
            'Transportes Andes SRL',
            '20456789123',
            'Av. Los Andes 789, Puno',
            '999888777 111222333 444555666',  # Múltiples teléfonos
            'andes@transportespuno.gob.pe',
            '11223344',
            'María Elena',
            'Quispe Mamani'
        ]
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws_datos.cell(row=1, column=col, value=header)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Escribir datos de ejemplo
    for row_idx, row_data in enumerate(datos_ejemplo, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_datos.cell(row=row_idx, column=col_idx, value=value)
            cell.font = texto_font
            cell.border = border
            if col_idx == 4:  # Columna teléfono
                cell.font = Font(name='Arial', size=10, bold=True, color='0066CC')
            elif col_idx == 5:  # Columna email
                cell.font = Font(name='Arial', size=10, bold=True, color='CC6600')
    
    # Ajustar ancho de columnas
    column_widths = [25, 15, 30, 25, 35, 15, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws_datos.column_dimensions[chr(64 + i)].width = width
    
    # Hoja para carga masiva (vacía)
    ws_carga = wb.create_sheet("CARGA_MASIVA")
    
    # Escribir solo encabezados
    for col, header in enumerate(headers, 1):
        cell = ws_carga.cell(row=1, column=col, value=header)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Ajustar ancho de columnas
    for i, width in enumerate(column_widths, 1):
        ws_carga.column_dimensions[chr(64 + i)].width = width
    
    # Agregar comentarios explicativos
    ws_carga['D1'].comment = "Puede ingresar múltiples números separados por espacios. Ej: 987654321 123456789"
    ws_carga['E1'].comment = "Debe usar @transportespuno.gob.pe"
    ws_carga['F1'].comment = "Obligatorio - 8 dígitos exactos"
    ws_carga['G1'].comment = "Opcional - se completa automáticamente si está vacío"
    ws_carga['H1'].comment = "Opcional - se completa automáticamente si está vacío"
    
    # Guardar archivo
    filename = 'plantilla_empresas_actualizada.xlsx'
    wb.save(filename)
    
    print(f"✅ Plantilla creada exitosamente: {filename}")
    print("\n📋 Características de la nueva plantilla:")
    print("   • Múltiples teléfonos separados por espacios")
    print("   • Representante con DNI obligatorio, nombres/apellidos opcionales")
    print("   • Email institucional @transportespuno.gob.pe")
    print("   • Validaciones automáticas implementadas")
    print("   • Ejemplos claros en hoja separada")
    
    return filename

if __name__ == "__main__":
    crear_plantilla_empresas_actualizada()