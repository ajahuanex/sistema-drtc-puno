#!/usr/bin/env python3
"""
Script para crear plantilla de resoluciones con ejemplos de años de vigencia
"""
import pandas as pd
from datetime import datetime

def crear_plantilla_con_ejemplos():
    """Crear plantilla Excel con ejemplos de resoluciones con diferentes años de vigencia"""
    
    # Datos de ejemplo con diferentes años de vigencia
    datos = {
        'Resolución Padre': ['', '', 'R-1001-2024', 'R-1002-2024'],
        'Número Resolución': ['1001-2024', '1002-2024', '1003-2024', '1004-2024'],
        'RUC Empresa': ['20123456789', '20234567890', '20123456789', '20234567890'],
        'Fecha Emisión': ['15/01/2024', '20/03/2024', '25/01/2024', '30/03/2024'],
        'Fecha Vigencia Inicio': ['15/01/2024', '20/03/2024', '', ''],
        'Años Vigencia': ['4', '10', '', ''],
        'Fecha Vigencia Fin': ['14/01/2028', '19/03/2034', '', ''],
        'Tipo Resolución': ['PADRE', 'PADRE', 'HIJO', 'HIJO'],
        'Tipo Trámite': ['PRIMIGENIA', 'PRIMIGENIA', 'RENOVACION', 'MODIFICACION'],
        'Descripción': [
            'Autorización para operar rutas interprovinciales - Vigencia 4 años',
            'Autorización para operar rutas interprovinciales - Vigencia 10 años',
            'Renovación de autorización de transporte',
            'Modificación de rutas autorizadas'
        ],
        'ID Expediente': ['123-2024', '456-2024', '789-2024', '012-2024'],
        'Usuario Emisión': ['USR001', 'USR001', 'USR001', 'USR001'],
        'Estado': ['VIGENTE', 'VIGENTE', 'VIGENTE', 'VIGENTE'],
        'Observaciones': [
            'Resolución padre con 4 años de vigencia',
            'Resolución padre con 10 años de vigencia',
            'Resolución hija - hereda vigencia del padre',
            'Resolución hija - hereda vigencia del padre'
        ]
    }
    
    df = pd.DataFrame(datos)
    
    # Crear archivo Excel
    nombre_archivo = f'plantilla_resoluciones_vigencia_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Resoluciones', index=False)
        
        # Obtener el workbook y worksheet para formatear
        workbook = writer.book
        worksheet = writer.sheets['Resoluciones']
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Agregar comentarios explicativos
        from openpyxl.comments import Comment
        
        worksheet['A1'].comment = Comment("Número de resolución padre (solo para resoluciones hijas). Dejar vacío para resoluciones padre.", "Sistema")
        worksheet['B1'].comment = Comment("Número sin prefijo R-. Ejemplo: 1001-2024 (el sistema agregará R-)", "Sistema")
        worksheet['D1'].comment = Comment("Formato: dd/mm/yyyy. Ejemplo: 15/01/2024", "Sistema")
        worksheet['E1'].comment = Comment("Solo para resoluciones PADRE. Fecha de inicio de vigencia.", "Sistema")
        worksheet['F1'].comment = Comment("Años de vigencia (4 o 10). Solo para resoluciones PADRE.", "Sistema")
        worksheet['G1'].comment = Comment("Se calcula automáticamente: Fecha Inicio + Años - 1 día. Puede dejarse vacío o usarse para validación.", "Sistema")
        worksheet['K1'].comment = Comment("OPCIONAL. Formatos aceptados: 123-2024, E-123-2024, 123-2024-E", "Sistema")
        
        # Colorear filas según tipo
        from openpyxl.styles import PatternFill
        
        # Color para resoluciones PADRE
        fill_padre = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        # Color para resoluciones HIJO
        fill_hijo = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
            tipo_resolucion = worksheet[f'H{row_idx}'].value
            fill = fill_padre if tipo_resolucion == 'PADRE' else fill_hijo
            for cell in row:
                cell.fill = fill
    
    print("=" * 70)
    print("✅ PLANTILLA CREADA EXITOSAMENTE")
    print("=" * 70)
    print(f"\n📄 Archivo: {nombre_archivo}")
    print("\n📋 Contenido:")
    print("-" * 70)
    print(df.to_string(index=False))
    print("-" * 70)
    
    print("\n💡 EJEMPLOS DE CÁLCULO:")
    print("-" * 70)
    print("Resolución 1001-2024:")
    print("  - Fecha Inicio: 15/01/2024")
    print("  - Años Vigencia: 4")
    print("  - Fecha Fin: 14/01/2028 (15/01/2024 + 4 años - 1 día)")
    print()
    print("Resolución 1002-2024:")
    print("  - Fecha Inicio: 20/03/2024")
    print("  - Años Vigencia: 10")
    print("  - Fecha Fin: 19/03/2034 (20/03/2024 + 10 años - 1 día)")
    print("-" * 70)
    
    print("\n📝 NOTAS:")
    print("  • Las resoluciones PADRE deben tener Fecha Inicio y Años Vigencia")
    print("  • Las resoluciones HIJO heredan la vigencia del padre")
    print("  • La Fecha Fin se calcula automáticamente")
    print("  • Si se proporciona Fecha Fin, se validará contra el cálculo")
    print("  • Años de vigencia típicos: 4 o 10 años")
    
    return nombre_archivo

if __name__ == "__main__":
    archivo = crear_plantilla_con_ejemplos()
    print(f"\n✅ Listo para usar: {archivo}")
