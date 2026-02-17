#!/usr/bin/env python3
"""
Script para generar plantilla de resoluciones con empresas reales de la BD
"""
import asyncio
import sys
import os
from datetime import datetime
import pandas as pd

# Agregar el directorio backend al path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

async def generar_plantilla_con_empresas_reales():
    """Generar plantilla con empresas reales de la base de datos"""
    from app.dependencies.db import get_database
    
    print("=" * 70)
    print("GENERANDO PLANTILLA CON EMPRESAS REALES")
    print("=" * 70)
    
    try:
        # Conectar a la base de datos
        print("\n⏳ Conectando a la base de datos...")
        db = await get_database()
        empresas_collection = db["empresas"]
        
        # Obtener empresas activas
        print("⏳ Obteniendo empresas activas...")
        empresas = await empresas_collection.find({
            "estaActivo": True
        }).limit(5).to_list(length=5)
        
        if not empresas:
            print("\n⚠️  No se encontraron empresas en la base de datos")
            print("   Usando datos de ejemplo...")
            empresas = [
                {"ruc": "20123456789", "razonSocial": {"principal": "Empresa de Ejemplo 1"}},
                {"ruc": "20234567890", "razonSocial": {"principal": "Empresa de Ejemplo 2"}}
            ]
        else:
            print(f"✅ Se encontraron {len(empresas)} empresas")
        
        # Crear datos para la plantilla
        datos = {
            'Resolución Padre': [],
            'Número Resolución': [],
            'RUC Empresa': [],
            'Fecha Emisión': [],
            'Fecha Vigencia Inicio': [],
            'Años Vigencia': [],
            'Fecha Vigencia Fin': [],
            'Tipo Resolución': [],
            'Tipo Trámite': [],
            'Descripción': [],
            'ID Expediente': [],
            'Usuario Emisión': [],
            'Estado': [],
            'Observaciones': []
        }
        
        # Generar resoluciones de ejemplo
        contador = 1001
        for i, empresa in enumerate(empresas[:2]):  # Solo 2 empresas para el ejemplo
            ruc = empresa.get('ruc', f'2012345678{i}')
            razon_social = empresa.get('razonSocial', {}).get('principal', f'Empresa {i+1}')
            
            # Resolución PADRE con 4 años
            datos['Resolución Padre'].append('')
            datos['Número Resolución'].append(f'{contador}-2024')
            datos['RUC Empresa'].append(ruc)
            datos['Fecha Emisión'].append('15/01/2024')
            datos['Fecha Vigencia Inicio'].append('15/01/2024')
            datos['Años Vigencia'].append('4')
            datos['Fecha Vigencia Fin'].append('14/01/2028')
            datos['Tipo Resolución'].append('PADRE')
            datos['Tipo Trámite'].append('PRIMIGENIA')
            datos['Descripción'].append(f'Autorización para operar rutas interprovinciales - {razon_social}')
            datos['ID Expediente'].append(f'{contador}-2024')
            datos['Usuario Emisión'].append('USR001')
            datos['Estado'].append('VIGENTE')
            datos['Observaciones'].append('Resolución padre con 4 años de vigencia')
            
            resolucion_padre_4 = f'R-{contador}-2024'
            contador += 1
            
            # Resolución PADRE con 10 años
            datos['Resolución Padre'].append('')
            datos['Número Resolución'].append(f'{contador}-2024')
            datos['RUC Empresa'].append(ruc)
            datos['Fecha Emisión'].append('20/03/2024')
            datos['Fecha Vigencia Inicio'].append('20/03/2024')
            datos['Años Vigencia'].append('10')
            datos['Fecha Vigencia Fin'].append('19/03/2034')
            datos['Tipo Resolución'].append('PADRE')
            datos['Tipo Trámite'].append('PRIMIGENIA')
            datos['Descripción'].append(f'Autorización para operar rutas interprovinciales - {razon_social}')
            datos['ID Expediente'].append(f'{contador}-2024')
            datos['Usuario Emisión'].append('USR001')
            datos['Estado'].append('VIGENTE')
            datos['Observaciones'].append('Resolución padre con 10 años de vigencia')
            
            resolucion_padre_10 = f'R-{contador}-2024'
            contador += 1
            
            # Resolución HIJO de la resolución con 4 años
            datos['Resolución Padre'].append(resolucion_padre_4)
            datos['Número Resolución'].append(f'{contador}-2024')
            datos['RUC Empresa'].append(ruc)
            datos['Fecha Emisión'].append('25/01/2024')
            datos['Fecha Vigencia Inicio'].append('')
            datos['Años Vigencia'].append('')
            datos['Fecha Vigencia Fin'].append('')
            datos['Tipo Resolución'].append('HIJO')
            datos['Tipo Trámite'].append('RENOVACION')
            datos['Descripción'].append(f'Renovación de autorización - {razon_social}')
            datos['ID Expediente'].append(f'{contador}-2024')
            datos['Usuario Emisión'].append('USR001')
            datos['Estado'].append('VIGENTE')
            datos['Observaciones'].append('Resolución hija - hereda vigencia del padre (4 años)')
            
            contador += 1
            
            # Resolución HIJO de la resolución con 10 años
            datos['Resolución Padre'].append(resolucion_padre_10)
            datos['Número Resolución'].append(f'{contador}-2024')
            datos['RUC Empresa'].append(ruc)
            datos['Fecha Emisión'].append('30/03/2024')
            datos['Fecha Vigencia Inicio'].append('')
            datos['Años Vigencia'].append('')
            datos['Fecha Vigencia Fin'].append('')
            datos['Tipo Resolución'].append('HIJO')
            datos['Tipo Trámite'].append('MODIFICACION')
            datos['Descripción'].append(f'Modificación de rutas autorizadas - {razon_social}')
            datos['ID Expediente'].append(f'{contador}-2024')
            datos['Usuario Emisión'].append('USR001')
            datos['Estado'].append('VIGENTE')
            datos['Observaciones'].append('Resolución hija - hereda vigencia del padre (10 años)')
            
            contador += 1
        
        # Crear DataFrame
        df = pd.DataFrame(datos)
        
        # Guardar en Excel
        nombre_archivo = f'plantilla_vigencia_empresas_reales_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Resoluciones', index=False)
            
            # Formatear
            workbook = writer.book
            worksheet = writer.sheets['Resoluciones']
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Agregar comentarios
            from openpyxl.comments import Comment
            
            worksheet['A1'].comment = Comment("Número de resolución padre (solo para resoluciones hijas)", "Sistema")
            worksheet['F1'].comment = Comment("Años de vigencia (4 o 10). Solo para resoluciones PADRE.", "Sistema")
            worksheet['G1'].comment = Comment("Se calcula automáticamente: Fecha Inicio + Años - 1 día", "Sistema")
            
            # Colorear filas
            from openpyxl.styles import PatternFill
            
            fill_padre = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            fill_hijo = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            
            for row_idx in range(2, len(df) + 2):
                tipo_resolucion = worksheet[f'H{row_idx}'].value
                fill = fill_padre if tipo_resolucion == 'PADRE' else fill_hijo
                for col_idx in range(1, 15):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill
        
        print("\n" + "=" * 70)
        print("✅ PLANTILLA GENERADA EXITOSAMENTE")
        print("=" * 70)
        print(f"\n📄 Archivo: {nombre_archivo}")
        print(f"\n📊 Estadísticas:")
        print(f"   Total de resoluciones: {len(df)}")
        print(f"   Resoluciones PADRE: {len(df[df['Tipo Resolución'] == 'PADRE'])}")
        print(f"   Resoluciones HIJO: {len(df[df['Tipo Resolución'] == 'HIJO'])}")
        print(f"   Con 4 años de vigencia: {len(df[df['Años Vigencia'] == '4'])}")
        print(f"   Con 10 años de vigencia: {len(df[df['Años Vigencia'] == '10'])}")
        
        print("\n📋 Empresas incluidas:")
        for i, empresa in enumerate(empresas[:2], 1):
            ruc = empresa.get('ruc', 'N/A')
            razon_social = empresa.get('razonSocial', {}).get('principal', 'N/A')
            print(f"   {i}. RUC: {ruc} - {razon_social}")
        
        print("\n💡 INSTRUCCIONES:")
        print("   1. Abrir el archivo Excel generado")
        print("   2. Revisar los datos de ejemplo")
        print("   3. Modificar según sea necesario")
        print("   4. Usar en: http://localhost:4200/resoluciones/carga-masiva")
        
        print("\n📝 NOTAS:")
        print("   • Las resoluciones PADRE tienen 'Años Vigencia' (4 o 10)")
        print("   • Las resoluciones HIJO heredan la vigencia del padre")
        print("   • La 'Fecha Vigencia Fin' se calcula automáticamente")
        print("   • Fórmula: Fecha Inicio + Años - 1 día")
        
        return nombre_archivo
        
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """Función principal"""
    print("\n🚀 Iniciando generación de plantilla...")
    archivo = asyncio.run(generar_plantilla_con_empresas_reales())
    
    if archivo:
        print(f"\n✅ Plantilla lista: {archivo}")
        return 0
    else:
        print("\n❌ Error al generar plantilla")
        return 1

if __name__ == "__main__":
    sys.exit(main())
